"""
Export service pre ILUMINATI SYSTEM
Podporuje Excel (xlsx), CSV, JSON export
"""

import io
import json
from datetime import datetime
from typing import Dict, List, Optional

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_excel(graph_data: Dict, filename: Optional[str] = None) -> bytes:
    """
    Exportuje grafové dáta do Excel (xlsx) formátu.

    Args:
        graph_data: Dict s nodes a edges
        filename: Voliteľný názov súboru

    Returns:
        bytes: Excel súbor ako bytes
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl nie je nainštalovaný. Nainštalujte: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Výsledky vyhľadávania"

    # Stylovanie hlavičky
    header_fill = PatternFill(
        start_color="0B4EA2", end_color="0B4EA2", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # === SHEET 1: Nodes (Firmy, Osoby, Adresy) ===
    ws.append(
        ["Typ", "ID", "Názov", "Krajina", "Risk Score", "Detaily", "Dátum vytvorenia"]
    )

    # Stylovanie hlavičky
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Pridať nodes
    nodes = graph_data.get("nodes", [])
    for node in nodes:
        ws.append(
            [
                node.get("type", ""),
                node.get("id", ""),
                node.get("label", ""),
                node.get("country", ""),
                node.get("risk_score", 0) or 0,
                json.dumps(node.get("details", {}), ensure_ascii=False)
                if isinstance(node.get("details"), dict)
                else str(node.get("details", "")),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    # Auto-width stĺpcov
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # === SHEET 2: Edges (Vzťahy) ===
    ws2 = wb.create_sheet("Vzťahy")
    ws2.append(["Source", "Target", "Typ vzťahu", "Váha"])

    # Stylovanie hlavičky
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Pridať edges
    edges = graph_data.get("edges", [])
    for edge in edges:
        ws2.append(
            [
                edge.get("source", ""),
                edge.get("target", ""),
                edge.get("type", ""),
                edge.get("weight", 1) or 1,
            ]
        )

    # Auto-width pre edges
    for column in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width

    # === SHEET 3: Summary (Súhrn) ===
    ws3 = wb.create_sheet("Súhrn")
    ws3.append(["Metrika", "Hodnota"])

    # Stylovanie hlavičky
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Pridať štatistiky
    summary_data = [
        ["Celkový počet nodov", len(nodes)],
        ["Celkový počet vzťahov", len(edges)],
        ["Dátum exportu", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Typy nodov", ""],
    ]

    # Počítanie typov nodov
    node_types = {}
    for node in nodes:
        node_type = node.get("type", "Unknown")
        node_types[node_type] = node_types.get(node_type, 0) + 1

    for node_type, count in sorted(node_types.items()):
        summary_data.append([node_type, count])

    summary_data.append(["", ""])
    summary_data.append(["Typy vzťahov", ""])

    # Počítanie typov vzťahov
    edge_types = {}
    for edge in edges:
        edge_type = edge.get("type", "Unknown")
        edge_types[edge_type] = edge_types.get(edge_type, 0) + 1

    for edge_type, count in sorted(edge_types.items()):
        summary_data.append([edge_type, count])

    for row in summary_data:
        ws3.append(row)

    # Auto-width pre summary
    for column in ws3.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws3.column_dimensions[column_letter].width = adjusted_width

    # Uložiť do bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_csv(graph_data: Dict) -> str:
    """
    Exportuje grafové dáta do CSV formátu.

    Args:
        graph_data: Dict s nodes a edges

    Returns:
        str: CSV obsah ako string
    """
    csv_lines = []

    # Nodes
    csv_lines.append("Typ,ID,Label,Krajina,Risk Score,Detaily")
    nodes = graph_data.get("nodes", [])
    for node in nodes:
        details = (
            json.dumps(node.get("details", {}), ensure_ascii=False)
            if isinstance(node.get("details"), dict)
            else str(node.get("details", ""))
        )
        csv_lines.append(
            f'{node.get("type", "")},{node.get("id", "")},"{node.get("label", "")}",{node.get("country", "")},{node.get("risk_score", 0) or 0},"{details.replace('"', '""')}"'
        )

    # Edges
    csv_lines.append("")
    csv_lines.append("Vzťahy:")
    csv_lines.append("Source,Target,Type,Weight")
    edges = graph_data.get("edges", [])
    for edge in edges:
        csv_lines.append(
            f"{edge.get('source', '')},{edge.get('target', '')},{edge.get('type', '')},{edge.get('weight', 1) or 1}"
        )

    return "\n".join(csv_lines)


def export_batch_to_excel(
    companies: List[Dict], filename: Optional[str] = None
) -> bytes:
    """
    Exportuje batch firiem do Excel (xlsx) formátu.

    Args:
        companies: List firiem (každá firma je dict s company_data)
        filename: Voliteľný názov súboru

    Returns:
        bytes: Excel súbor ako bytes
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl nie je nainštalovaný. Nainštalujte: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Export"

    # Stylovanie hlavičky
    header_fill = PatternFill(
        start_color="0B4EA2", end_color="0B4EA2", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Hlavička
    headers = [
        "IČO/KRS/Adószám",
        "Názov",
        "Krajina",
        "Adresa",
        "DIČ",
        "IČ DPH",
        "Právna forma",
        "Risk Score",
        "Dátum vzniku",
        "Stav",
        "Poznámky",
    ]
    ws.append(headers)

    # Stylovanie hlavičky
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Pridať firmy
    for company in companies:
        company_data = company.get("company_data", {}) or company
        ws.append(
            [
                company_data.get("ico") or company.get("company_identifier", ""),
                company_data.get("name") or company.get("company_name", ""),
                company_data.get("country") or company.get("country", ""),
                company_data.get("address", ""),
                company_data.get("dic", ""),
                company_data.get("ic_dph", ""),
                company_data.get("legal_form", ""),
                company.get("risk_score") or company_data.get("risk_score", 0),
                company_data.get("establishment_date", ""),
                company_data.get("status", ""),
                company.get("notes", ""),
            ]
        )

    # Auto-width stĺpcov
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Uložiť do bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
