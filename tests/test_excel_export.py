"""
Špeciálne testy pre Excel export funkcionalitu
"""

import io

import pytest
import requests
import urllib3

# Skúsiť importovať openpyxl
try:
    from openpyxl import load_workbook  # type: ignore[reportMissingModuleSource]

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None

# Potlač SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL = "http://localhost:8000"
BASE_URL_HTTPS = "https://localhost:8000"


def get_base_url():
    """Vráti dostupný base URL"""
    try:
        requests.get(f"{BASE_URL_HTTPS}/api/health", verify=False, timeout=2)
        return BASE_URL_HTTPS
    except Exception:
        try:
            requests.get(f"{BASE_URL}/api/health", timeout=2)
            return BASE_URL
        except Exception:
            pytest.skip("Backend server nie je dostupný")


def test_excel_export_endpoint_exists():
    """Test, či Excel export endpoint existuje"""
    base_url = get_base_url()
    verify_ssl = base_url.startswith("https")

    # Test, či endpoint existuje (bez autentifikácie by mal vrátiť 403)
    try:
        response = requests.post(
            f"{base_url}/api/export/excel",
            json={"nodes": [], "edges": []},
            verify=verify_ssl,
            timeout=5,
        )
        # Endpoint by mal existovať (403 = forbidden, 404 = not found)
        assert response.status_code != 404, "Excel export endpoint neexistuje"
    except requests.exceptions.ConnectionError:
        pytest.skip("Backend server nie je dostupný")


def test_batch_excel_export_endpoint_exists():
    """Test, či batch Excel export endpoint existuje"""
    base_url = get_base_url()
    verify_ssl = base_url.startswith("https")

    try:
        response = requests.post(
            f"{base_url}/api/export/batch-excel", json=[], verify=verify_ssl, timeout=5
        )
        # Endpoint by mal existovať
        assert response.status_code != 404, "Batch Excel export endpoint neexistuje"
    except requests.exceptions.ConnectionError:
        pytest.skip("Backend server nie je dostupný")


def test_excel_export_service_available():
    """Test, či Excel export service je dostupný (openpyxl)"""
    try:
        from backend.services.export_service import (
            export_batch_to_excel,
            export_to_excel,
        )

        assert export_to_excel is not None
        assert export_batch_to_excel is not None
    except ImportError:
        pytest.skip(
            "Excel export service nie je dostupný (openpyxl nie je nainštalovaný)"
        )


def test_excel_export_creates_valid_file():
    """Test, či Excel export vytvára platný Excel súbor"""
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl nie je nainštalovaný")

    try:
        from backend.services.export_service import export_to_excel

        # Test dáta
        test_data = {
            "nodes": [
                {
                    "id": "test_1",
                    "label": "Test Firma",
                    "type": "company",
                    "country": "SK",
                    "risk_score": 5,
                    "details": {"ico": "12345678"},
                }
            ],
            "edges": [],
        }

        # Exportovať do Excel
        excel_bytes = export_to_excel(test_data)

        # Skontrolovať, či je to platný Excel súbor
        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # Skúsiť načítať pomocou openpyxl
        if load_workbook is None:
            pytest.skip("openpyxl nie je nainštalovaný")
        workbook = load_workbook(io.BytesIO(excel_bytes))
        assert workbook is not None
        assert len(workbook.sheetnames) > 0

        # Skontrolovať, či obsahuje očakávané sheets
        assert "Výsledky vyhľadávania" in workbook.sheetnames or any(
            "výsledky" in name.lower() for name in workbook.sheetnames
        )

    except ImportError:
        pytest.skip("openpyxl nie je nainštalovaný")
    except Exception as e:
        pytest.fail(f"Excel export zlyhal: {e}")


def test_batch_excel_export_creates_valid_file():
    """Test, či batch Excel export vytvára platný Excel súbor"""
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl nie je nainštalovaný")

    try:
        from backend.services.export_service import export_batch_to_excel

        # Test dáta
        test_companies = [
            {
                "company_identifier": "12345678",
                "company_name": "Test Firma 1",
                "country": "SK",
                "risk_score": 5,
                "company_data": {"ico": "12345678", "name": "Test Firma 1"},
            },
            {
                "company_identifier": "87654321",
                "company_name": "Test Firma 2",
                "country": "CZ",
                "risk_score": 3,
                "company_data": {"ico": "87654321", "name": "Test Firma 2"},
            },
        ]

        # Exportovať do Excel
        excel_bytes = export_batch_to_excel(test_companies)

        # Skontrolovať, či je to platný Excel súbor
        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # Skúsiť načítať pomocou openpyxl
        if load_workbook is None:
            pytest.skip("openpyxl nie je nainštalovaný")
        workbook = load_workbook(io.BytesIO(excel_bytes))
        assert workbook is not None
        assert len(workbook.sheetnames) > 0

    except ImportError:
        pytest.skip("openpyxl nie je nainštalovaný")
    except Exception as e:
        pytest.fail(f"Batch Excel export zlyhal: {e}")
